"""
ingest_xlsx.py — Excel (.xlsx) to Markdown converter for process-ai ingest pipeline.

Converts Excel workbooks to markdown tables. Uses openpyxl (data_only mode
to read computed values, not formulas). Each worksheet becomes a markdown
section with a pipe table.
"""

import argparse
import os
import sys

# Import shared helpers from sibling script (same directory)
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from ingest_common import (
    emit_error, emit_success, extract_title_from_md,
    sanitize_text, slugify_for_path,
)

MAX_ROWS_PER_SHEET = 500


def _sheet_to_markdown(ws, sheet_name: str) -> str:
    """Convert a single worksheet to markdown section."""
    lines = [f"## {sheet_name}\n"]

    # Collect all rows (data_only already applied at workbook level)
    rows = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        # Skip completely empty rows
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        if len(rows) >= MAX_ROWS_PER_SHEET:
            break
        rows.append(row)

    if not rows:
        lines.append(f"*(planilha vazia)*\n")
        return '\n'.join(lines)

    # Determine max columns from data
    max_cols = max(len(row) for row in rows) if rows else 0
    if max_cols == 0:
        lines.append(f"*(planilha vazia)*\n")
        return '\n'.join(lines)

    # First row → header
    header = rows[0]
    data_rows = rows[1:] if len(rows) > 1 else []

    padded_header = list(header) + [''] * (max_cols - len(header))
    lines.append('| ' + ' | '.join(sanitize_text(str(c)) or ' ' for c in padded_header) + ' |')
    lines.append('| ' + ' | '.join('---' for _ in range(max_cols)) + ' |')

    for row in data_rows:
        padded_row = list(row) + [''] * (max_cols - len(row))
        cells = []
        for c in padded_row:
            val = c if c is not None else ''
            cells.append(sanitize_text(str(val)) or ' ')
        lines.append('| ' + ' | '.join(cells) + ' |')

    if len(ws.rows) > MAX_ROWS_PER_SHEET:
        lines.append(f'\n*(planilha truncada — {MAX_ROWS_PER_SHEET} de mais linhas exibidas)*')

    return '\n'.join(lines)


def convert_xlsx(xlsx_path: str, output_dir: str):
    """Convert an Excel workbook to Markdown. Returns (md_path, images, metadata, sheets) or (None, [], {}, 0) on error."""
    # openpyxl is imported here so the import error is only raised when
    # actually processing .xlsx files, not on module load.
    try:
        from openpyxl import load_workbook
    except ImportError:
        emit_error(
            "Biblioteca openpyxl não encontrada. "
            "Instale as dependências: pip install -r scripts/requirements-ingest.txt"
        )
        return None, [], {}, 0

    filename = os.path.basename(xlsx_path)
    base_name = os.path.splitext(filename)[0]
    slug = slugify_for_path(base_name)

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    # Metadata from workbook properties
    author = ''
    title_prop = ''
    created = ''
    try:
        if wb.properties:
            author = str(wb.properties.creator or '')
            title_prop = str(wb.properties.title or '')
            created = str(wb.properties.created or '')
    except Exception:
        pass

    sections = [f"# {title_prop or base_name}\n"]

    sheet_count = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        section = _sheet_to_markdown(ws, sheet_name)
        if sheet_count > 0:
            sections.append('---')
        sections.append(section)
        sheet_count += 1

    wb.close()

    md_body = '\n\n'.join(sections)

    # Write markdown
    md_name = f"{slug}.md"
    md_path = os.path.join(output_dir, md_name)
    with open(md_path, 'w', encoding='utf-8') as f:
        f.write(md_body)

    title = extract_title_from_md(md_body) or title_prop or base_name
    metadata = {
        'source_file': filename,
        'title': title,
        'author': author,
        'created': created,
    }

    return md_path, [], metadata, sheet_count


def main():
    parser = argparse.ArgumentParser(description='Convert XLSX to Markdown')
    parser.add_argument('--input', required=True, help='Path to XLSX file')
    parser.add_argument('--output-dir', required=True,
                        help='Directory for output markdown')
    args = parser.parse_args()

    xlsx_path = os.path.abspath(args.input)
    output_dir = os.path.abspath(args.output_dir)

    if not os.path.isfile(xlsx_path):
        emit_error(f"Arquivo não encontrado: {xlsx_path}")
        return

    try:
        os.makedirs(output_dir, exist_ok=True)
        md_path, images, metadata, sheets = convert_xlsx(xlsx_path, output_dir)
        if md_path is None:
            return  # error already emitted via emit_error
        md_rel = os.path.relpath(md_path, output_dir).replace('\\', '/')

        emit_success(
            format='xlsx',
            markdown_rel=md_rel,
            images=images,
            metadata=metadata,
            sheets=sheets,
        )
    except Exception as e:
        emit_error(f"Falha ao converter XLSX '{os.path.basename(xlsx_path)}': {str(e)}")


if __name__ == "__main__":
    main()
